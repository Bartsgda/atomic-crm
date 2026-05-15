"""
link_sub_agents.py — linkowanie policy_sub_agent_shares z XLSX col[13].
Spec: src/legacy-v1/AUDIT_PLAN.md § Faza 1.5 (Sub_agent linking)
Uruchomiono po: scripts/fix_audit_rows.py (Faza 1).

ZRODLO: BAZA_bez_pesel.xlsx col[13] = 'kogo' (nie ai_parsed_182.json — ten JSON
        zawiera dane pojazdu/polisy, NIE sub_agent linking).

KOLUMNA col[13] (kogo) — reguly parsowania (z IMPORT_AUDIT_HOWTO.md § 4):
  'wlasny' / 'własny'            -> sub_agents WHERE group_prefix='wlasny' (match: 'Własny')
  'firmowy'                      -> sub_agents WHERE name='Firmowy (ogólny)' group_prefix='firmowy'
  'firmowy/Beata'                -> name='Beata' group_prefix='firmowy'
  'firmowy/Osip/Ewa W'           -> AMBIGUOUS: 2 nazwiska, skip + raport
  'Imie Nazwisko'                -> name match (zewnetrzny partner)
  null / '' / '?'                -> brak sub_agent, skip

Logika dopasowania sub_agent.name:
  1. Exact match (case-insensitive)
  2. Prefix match (np. 'Osip' matchuje 'Osip Jakub Alot')
  3. Fuzzy: pierwsze slowo match (dla jednowyrazowych wartosci jak 'Ania', 'Beata')

Usage:
  python scripts/link_sub_agents.py --dry-run        # domyslne: wyswietl tylko diff
  python scripts/link_sub_agents.py --apply          # INSERT do policy_sub_agent_shares
  python scripts/link_sub_agents.py --apply --xlsx-path "C:/inny/plik.xlsx"

PIASKOWNICA: schema=test. Idempotentne (nie duplikuje istniejacych par policy+sub_agent).
"""
import os
import sys
import json
import argparse
import urllib.request
import urllib.error
import urllib.parse
import re

# ============================================================
# KONFIGURACJA
# ============================================================

URL     = os.environ['CRM_ALINA_SUPABASE_URL'].rstrip('/')
KEY     = os.environ['CRM_ALINA_SB_SECRET']
SCHEMA  = 'test'
TENANT_ID = '11111111-1111-1111-1111-111111111111'

XLSX_DEFAULT = r'C:/BartsGda4/CRM-ALINA/DANE-POZNIEJ-USUN/BAZA_bez_pesel.xlsx'
XLSX_SHEET   = 'potencjalny'
# XLSX: header=wiersz 2, dane od wiersza 3 = legacy_id row_1
XLSX_DATA_START_ROW = 3
XLSX_COL_KOGO = 13   # col[13] = kolumna 14 (1-based)

IMPORT_NOTE = 'Import XLSX (lead-only) — link_sub_agents.py 2026-05-15'

# ============================================================
# HTTP HELPER
# ============================================================

def req(method, path, body=None):
    data = json.dumps(body).encode('utf-8') if body is not None else None
    headers = {
        'apikey':          KEY,
        'Authorization':   f'Bearer {KEY}',
        'Accept-Profile':  SCHEMA,
        'Content-Profile': SCHEMA,
        'Content-Type':    'application/json',
        'Prefer':          'return=representation',
        'User-Agent':      'postgrest-cli',
    }
    r = urllib.request.Request(
        f'{URL}/rest/v1/{path}',
        data=data, method=method, headers=headers)
    try:
        resp = urllib.request.urlopen(r)
        raw = resp.read()
        return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        err_body = e.read()[:400]
        print(f'  ERR {method} /rest/v1/{path}: HTTP {e.code} {err_body}')
        raise


def get(path):              return req('GET',  path)
def post_row(path, body):   return req('POST', path, body)


# ============================================================
# PARSOWANIE col[13] → sub_agent_id
# ============================================================

def parse_kogo(raw, name_to_sa, prefix_to_sa):
    """
    Parsuje surowa wartosc col[13] i zwraca (sa_id, info, is_ambiguous).
    sa_id=None gdy brak pasujacego agenta lub wartosc pusta/null.
    is_ambiguous=True gdy format firmowy/X/Y (wiele nazwisk).
    """
    if not raw:
        return None, 'brak wartosci col[13]', False

    val = str(raw).strip()
    if not val or val in ('?', '-', 'brak', 'n/d'):
        return None, f'wartosc ignorowana: {val!r}', False

    val_lower = val.lower()

    # wlasny/własny -> group_prefix='wlasny'
    # Obsluguje tez 'własny->rozliczona po zerwaniu...' (notatka po ->)
    val_for_wlasny = val_lower.split('->')[0].strip()  # odrzuc sufiks z '->'
    if val_for_wlasny in ('wlasny', 'własny'):
        sa_id = prefix_to_sa.get('wlasny')
        if sa_id:
            suffix_note = val[val.find('->'):] if '->' in val else ''
            info = f'wlasny -> {sa_id[:8]}' + (f' [sufiks: {suffix_note!r}]' if suffix_note else '')
            return sa_id, info, False
        return None, 'BRAK sub_agent z group_prefix=wlasny w DB', False

    # firmowy (sam) -> 'Firmowy (ogolny)'
    if val_lower == 'firmowy':
        sa_id = name_to_sa.get('Firmowy (ogólny)')
        if sa_id:
            return sa_id, 'firmowy ogolny', False
        return None, 'BRAK "Firmowy (ogólny)" w DB', False

    # firmowy/X[/Y...] -> match po czesci bez 'firmowy/' prefiksu
    # Format 'firmowy/Osip/Jakub Alot' -> proba match na 'Osip/Jakub Alot' (slash jako czesc nazwy)
    # przed probka 'Osip' (krotszy prefix). Iterujemy od najdluzszego do najkrotszego.
    if val_lower.startswith('firmowy/'):
        after_prefix = val[len('firmowy/'):].strip()  # 'Osip/Jakub Alot' lub 'Beata' etc.
        parts = after_prefix.split('/')

        # Proba: cala czesc po 'firmowy/' jako nazwa (np. 'Osip/Jakub Alot')
        sa_id = _match_name(after_prefix, name_to_sa)
        if sa_id:
            return sa_id, f'firmowy/{after_prefix}', False

        # Jesli wiele tokenow po /, proba na 'X/Y', 'X' itd. (od najdluzszego)
        for i in range(len(parts), 0, -1):
            candidate = '/'.join(parts[:i]).strip()
            sa_id = _match_name(candidate, name_to_sa)
            if sa_id:
                return sa_id, f'firmowy/{candidate} (prefix match z {after_prefix!r})', False

        # Podwojna spacja lub literowka: normalizuj spacje i proba
        normalized = re.sub(r'\s+', ' ', after_prefix).strip()
        if normalized != after_prefix:
            sa_id = _match_name(normalized, name_to_sa)
            if sa_id:
                return sa_id, f'firmowy/{normalized} (po normalizacji spacji)', False
            for i in range(len(normalized.split('/')), 0, -1):
                candidate = '/'.join(normalized.split('/')[:i]).strip()
                sa_id = _match_name(candidate, name_to_sa)
                if sa_id:
                    return sa_id, f'firmowy/{candidate} (prefix+normalizacja z {after_prefix!r})', False

        return None, f'BRAK match dla firmowy/{after_prefix!r}', False

    # Exact match (zewnetrzny partner lub specjalne nazwy)
    sa_id = _match_name(val, name_to_sa)
    if sa_id:
        return sa_id, f'exact/prefix match: {val!r}', False

    return None, f'BRAK match dla: {val!r}', False


def _match_name(query, name_to_sa):
    """Exact (case-insensitive), potem prefix match po pierwszym slowie."""
    q = query.strip()
    q_lower = q.lower()
    # Exact
    for name, sid in name_to_sa.items():
        if name.lower() == q_lower:
            return sid
    # Prefix match (nazwa w DB zaczyna sie od query)
    for name, sid in name_to_sa.items():
        if name.lower().startswith(q_lower + '/') or name.lower().startswith(q_lower + ' '):
            return sid
    # Single-word fuzzy: pierwszy token match
    tokens = q_lower.split()
    if len(tokens) == 1:
        for name, sid in name_to_sa.items():
            if name.lower().split()[0] == tokens[0]:
                return sid
    return None


# ============================================================
# MAIN
# ============================================================

def main():
    ap = argparse.ArgumentParser(
        description='Linkuj policy_sub_agent_shares z XLSX col[13]. Dry-run domyslnie.'
    )
    ap.add_argument('--dry-run', action='store_true', help='tylko wyswietl diff (domyslne bez flag)')
    ap.add_argument('--apply',   action='store_true', help='faktyczny INSERT do policy_sub_agent_shares')
    ap.add_argument('--xlsx-path', default=XLSX_DEFAULT, help='sciezka do XLSX (domyslnie BAZA_bez_pesel.xlsx)')
    args = ap.parse_args()

    if not args.dry_run and not args.apply:
        print('Wybierz --dry-run lub --apply')
        sys.exit(1)

    dry = not args.apply

    print(f'\n{"=" * 60}')
    print(f'LINK SUB AGENTS — {"DRY-RUN" if dry else "APPLY"}')
    print(f'Zrodlo XLSX: {args.xlsx_path}')
    print(f'Schema: {SCHEMA} | URL: {URL[:40]}...')
    print(f'{"=" * 60}')

    # Wczytaj XLSX
    try:
        import openpyxl
    except ImportError:
        print('BLAD: openpyxl nie zainstalowany. pip install openpyxl')
        sys.exit(3)

    try:
        wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)
        ws = wb[XLSX_SHEET]
    except FileNotFoundError:
        print(f'BLAD: plik XLSX nie znaleziony: {args.xlsx_path}')
        sys.exit(4)
    except KeyError:
        print(f'BLAD: arkusz "{XLSX_SHEET}" nie istnieje w pliku XLSX')
        sys.exit(4)

    # Zaladuj sub_agents z DB
    sub_agents_db = get('sub_agents?select=id,name,group_prefix,tenant_id')
    name_to_sa    = {sa['name']: sa['id'] for sa in sub_agents_db}
    prefix_to_sa  = {}
    for sa in sub_agents_db:
        gp = sa.get('group_prefix')
        if gp and gp not in prefix_to_sa:  # pierwszy match per prefix (wlasny -> 'Wlasny')
            prefix_to_sa[gp] = sa['id']

    print(f'  Sub_agents w DB: {len(sub_agents_db)}')
    print(f'  Prefix map: {list(prefix_to_sa.keys())}')

    # Zaladuj istniejace shares (dedup)
    existing_shares = get('policy_sub_agent_shares?select=policy_id,sub_agent_id')
    existing_set    = {(s['policy_id'], s['sub_agent_id']) for s in existing_shares}
    print(f'  Istniejace shares w DB: {len(existing_set)}')

    # Zaladuj wszystkie polisy (batch, zamiast N per-wiersz zapytan)
    all_policies = get('policies?select=id,legacy_id,tenant_id')
    legacy_to_policy = {p['legacy_id']: p for p in all_policies}
    print(f'  Polisy w DB: {len(all_policies)}')

    print()

    # Iteruj przez XLSX
    inserted     = 0
    skipped_none = 0
    skipped_ambiguous = []
    skipped_no_match  = []
    skipped_exists    = 0
    skipped_no_policy = []

    for r in range(XLSX_DATA_START_ROW, ws.max_row + 1):
        row_idx   = r - XLSX_DATA_START_ROW + 1   # 1-based legacy index
        legacy_id = f'xlsx_2025_row_{row_idx}'
        kogo_raw  = ws.cell(row=r, column=XLSX_COL_KOGO + 1).value  # openpyxl 1-based

        sa_id, info, is_ambiguous = parse_kogo(kogo_raw, name_to_sa, prefix_to_sa)

        if kogo_raw is None or (isinstance(kogo_raw, str) and not kogo_raw.strip()):
            skipped_none += 1
            continue

        if is_ambiguous:
            skipped_ambiguous.append(f'{legacy_id}: col[13]={kogo_raw!r} — {info}')
            continue

        if sa_id is None:
            if 'ignorowana' not in info and 'BRAK' not in info:
                skipped_none += 1
            else:
                skipped_no_match.append(f'{legacy_id}: col[13]={kogo_raw!r} — {info}')
            continue

        # Znajdz policy_id
        policy = legacy_to_policy.get(legacy_id)
        if not policy:
            skipped_no_policy.append(f'{legacy_id}: brak polisy w DB')
            continue

        policy_id = policy['id']
        tenant_id = policy.get('tenant_id') or TENANT_ID

        if (policy_id, sa_id) in existing_set:
            skipped_exists += 1
            continue

        sa_name = next((sa['name'] for sa in sub_agents_db if sa['id'] == sa_id), '?')
        print(f'  INSERT: {legacy_id} col13={kogo_raw!r} -> sub_agent "{sa_name}" ({info})')

        if not dry:
            post_row('policy_sub_agent_shares', {
                'policy_id':    policy_id,
                'sub_agent_id': sa_id,
                'tenant_id':    tenant_id,
                'rate':         None,
                'amount':       0,
                'note':         IMPORT_NOTE,
            })
            existing_set.add((policy_id, sa_id))
        inserted += 1

    print(f'\n{"=" * 60}')
    print('PODSUMOWANIE:')
    print(f'  Do insertu:              {inserted}')
    print(f'  Juz istniejace (skip):   {skipped_exists}')
    print(f'  Brak wartosci col[13]:   {skipped_none}')
    print(f'  Brak polisy w DB:        {len(skipped_no_policy)}')
    print(f'  Brak match sub_agent:    {len(skipped_no_match)}')
    print(f'  Ambiguous (wiele nazwisk): {len(skipped_ambiguous)}')

    if skipped_no_match:
        print('\n  Wiersze bez match sub_agent:')
        for s in skipped_no_match:
            print(f'    {s}')

    if skipped_ambiguous:
        print('\n  Ambiguous (wymagaja recznego sprawdzenia):')
        for s in skipped_ambiguous:
            print(f'    {s}')

    if skipped_no_policy:
        print('\n  Brakujace polisy w DB:')
        for s in skipped_no_policy[:10]:
            print(f'    {s}')

    if dry:
        print('\n  [DRY-RUN] Uruchom z --apply zeby zaaplikowac.')
    else:
        print(f'\n  Zaaplikowano {inserted} INSERT-ow do policy_sub_agent_shares.')

    print(f'{"=" * 60}')
    print(f'DONE ({"DRY-RUN" if dry else "APPLY"})')
    print(f'{"=" * 60}\n')


if __name__ == '__main__':
    main()
